#!/usr/bin/env python3
"""
Script de xuat bao cao loi nhuan ra file Excel voi 4 sheet:
1. Sheet 1: Tong loi nhuan - hien thi tong doanh thu va chi phi
2. Sheet 2: Doanh thu - chi tiet tung hoa don, san pham, khach hang
3. Sheet 3: Chi phi - chi tiet tung khoan phi
4. Sheet 4: Chi phi nhan su - chi tiet luong tung nhan vien

Tat ca duoc xuat theo thoi gian da chon
"""

import os
import sys
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Them thu muc backend vao path de import cac module
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

# Import config
try:
    from supabase_client import supabase, SUPABASE_AVAILABLE
    if not SUPABASE_AVAILABLE:
        raise ImportError("Supabase not available")
except ImportError:
    # Fallback neu khong co supabase_client
    from dotenv import load_dotenv
    from supabase import create_client, Client

    load_dotenv()
    SUPABASE_URL = os.getenv('SUPABASE_URL')
    SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_KEY') or os.getenv('SUPABASE_ANON_KEY')

    if not SUPABASE_URL or not SUPABASE_KEY:
        raise ValueError("Supabase credentials not found")

    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

def get_profit_data(month=None):
    """Lay du lieu loi nhuan tu database"""
    try:
        # Lay du lieu doanh thu
        revenue_query = supabase.table('invoices_reality').select('*')
        if month:
            start_date = f"{month}-01"
            year, month_num = map(int, month.split('-'))
            if month_num == 12:
                end_date = f"{year + 1}-01-01"
            else:
                end_date = f"{year}-{month_num + 1:02d}-01"
            revenue_query = revenue_query.gte('invoice_date', start_date).lt('invoice_date', end_date)

        revenue_result = revenue_query.execute()
        revenue_data = revenue_result.data

        # Lay du lieu chi phi
        expense_query = supabase.table('quanly_chiphi').select('*')
        if month:
            expense_query = expense_query.gte('created_at', start_date).lt('created_at', end_date)

        expense_result = expense_query.execute()
        expense_data = expense_result.data

        # Lay thong tin loai chi phi
        expense_category_ids = list(set(item.get('id_lcp') for item in expense_data if item.get('id_lcp')))
        expense_categories = {}
        if expense_category_ids:
            try:
                categories_result = supabase.table('loaichiphi').select('*').in_('id', expense_category_ids).execute()
                expense_categories = {item['id']: item for item in categories_result.data}
            except Exception as e:
                print(f"Error getting expense categories: {e}")

        # Lay chi tiet san pham cho hoa don
        invoice_ids = [inv['id'] for inv in revenue_data]
        invoice_items = []
        if invoice_ids:
            try:
                items_result = supabase.table('invoice_items_reality').select('*').in_('invoice_id', invoice_ids).execute()
                invoice_items = items_result.data
            except Exception as e:
                print(f"Error getting invoice items: {e}")

        # Lay thong tin lookup cho cac loai
        try:
            loainhom_result = supabase.table('loainhom').select('*').execute()
            loainhom_map = {str(item['id']): item['tenloai'] for item in loainhom_result.data}

            loaikinh_result = supabase.table('loaikinh').select('*').execute()
            loaikinh_map = {str(item['id']): item['tenloai'] for item in loaikinh_result.data}

            loaitaynam_result = supabase.table('loaitaynam').select('*').execute()
            loaitaynam_map = {str(item['id']): item['tenloai'] for item in loaitaynam_result.data}

            bophan_result = supabase.table('bophan').select('*').execute()
            bophan_map = {str(item['id']): item['tenloai'] for item in bophan_result.data}

            sanpham_result = supabase.table('sanpham').select('*').execute()
            sanpham_map = {}
            for item in sanpham_result.data:
                key = f"{item['id_nhom']}_{item['id_kinh']}_{item['id_taynam']}_{item['id_bophan']}"
                sanpham_map[key] = item['tensp']

        except Exception as e:
            print(f"Error getting lookup data: {e}")
            loainhom_map = {}
            loaikinh_map = {}
            loaitaynam_map = {}
            bophan_map = {}
            sanpham_map = {}

        # Lay du lieu chi phi nhan su (phieu luong)
        payroll_query = supabase.table('phieu_luong').select('*')
        if month:
            payroll_query = payroll_query.eq('ky_tinh_luong', month)

        payroll_result = payroll_query.execute()
        payroll_data = payroll_result.data

        # Lay thong tin nhan vien cho phieu luong
        employee_ids = list(set(item.get('ma_nv') for item in payroll_data if item.get('ma_nv')))
        employees = {}
        if employee_ids:
            try:
                employees_result = supabase.table('employees').select('*').in_('ma_nv', employee_ids).execute()
                employees = {emp['ma_nv']: emp for emp in employees_result.data}
            except Exception as e:
                print(f"Error getting employee data: {e}")

        return {
            'revenue': revenue_data,
            'expenses': expense_data,
            'payroll': payroll_data,
            'employees': employees,
            'expense_categories': expense_categories,
            'invoice_items': invoice_items,
            'month': month,
            'loainhom_map': loainhom_map,
            'loaikinh_map': loaikinh_map,
            'loaitaynam_map': loaitaynam_map,
            'bophan_map': bophan_map,
            'sanpham_map': sanpham_map
        }

    except Exception as e:
        print(f"Error getting data: {e}")
        return None

def create_excel_file(data, output_path):
    """Tao file Excel voi 3 sheet"""

    # Tao workbook
    wb = Workbook()

    # Xoa sheet mac dinh
    wb.remove(wb.active)

    # Tao 4 sheet
    sheet1 = wb.create_sheet("Tong loi nhuan")
    sheet2 = wb.create_sheet("Doanh thu chi tiet")
    sheet3 = wb.create_sheet("Chi phi chi tiet")
    sheet4 = wb.create_sheet("Chi phi nhan su")

    # Dinh dang chung
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== SHEET 1: TONG LOI NHUAN =====
    create_profit_summary_sheet(sheet1, data, header_font, header_fill, border)

    # ===== SHEET 2: DOANH THU CHI TIET =====
    create_revenue_detail_sheet(sheet2, data, header_font, header_fill, border)

    # ===== SHEET 3: CHI PHI CHI TIET =====
    create_expense_detail_sheet(sheet3, data, header_font, header_fill, border)

    # ===== SHEET 4: CHI PHI NHAN SU =====
    create_payroll_detail_sheet(sheet4, data, header_font, header_fill, border)

    # Luu file
    wb.save(output_path)
    return output_path

def create_profit_summary_sheet(sheet, data, header_font, header_fill, border):
    """Tao sheet tong loi nhuan"""

    # Tieu de
    sheet['A1'] = "BAO CAO TONG LOI NHUAN"
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A2'] = f"Thang: {data['month'] or 'Tat ca'}"
    sheet['A2'].font = Font(bold=True, size=12)

    # Tinh tong
    total_revenue = sum(inv.get('total_amount', 0) for inv in data['revenue'])
    total_expenses = sum(exp.get('giathanh', 0) for exp in data['expenses'])
    total_payroll = sum(p.get('luong_thuc_nhan', 0) for p in data['payroll'])
    total_all_expenses = total_expenses + total_payroll
    total_profit = total_revenue - total_all_expenses
    profit_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0

    # Header
    headers = ['Chi tieu', 'So tien (VND)', 'Ty le (%)']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Du lieu
    summary_data = [
        ['Tong doanh thu', total_revenue, '100.00'],
        ['Tong chi phi van hanh', total_expenses, f"{(total_expenses/total_revenue*100):.2f}" if total_revenue > 0 else '0.00'],
        ['Tong chi phi nhan su', total_payroll, f"{(total_payroll/total_revenue*100):.2f}" if total_revenue > 0 else '0.00'],
        ['Tong chi phi', total_all_expenses, f"{(total_all_expenses/total_revenue*100):.2f}" if total_revenue > 0 else '0.00'],
        ['Hoat dong kinh doanh', total_profit, f"{profit_margin:.2f}"],
    ]

    for row_idx, row_data in enumerate(summary_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if col_idx == 1:
                cell.value = value
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.value = value
                cell.alignment = Alignment(horizontal='right')
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
            cell.border = border

    # Them thong tin thong ke
    sheet['A10'] = "THONG KE CHI TIET"
    sheet['A10'].font = Font(bold=True, size=12)

    stats_data = [
        ['So luong hoa don', len(data['revenue'])],
        ['So luong chi phi van hanh', len(data['expenses'])],
        ['So luong nhan vien co luong', len(data['payroll'])],
        ['Trung binh luong/nhan vien', total_payroll / len(data['payroll']) if data['payroll'] else 0],
        ['Trang thai', 'HOAT DONG KINH DOANH TOT' if total_profit >= 0 else 'HOAT DONG KINH DOANH XAU'],
    ]

    for row_idx, (label, value) in enumerate(stats_data, 11):
        sheet.cell(row=row_idx, column=1).value = label
        sheet.cell(row=row_idx, column=1).font = Font(bold=True)
        sheet.cell(row=row_idx, column=1).border = border

        cell = sheet.cell(row=row_idx, column=2)
        cell.value = value
        cell.border = border
        if isinstance(value, (int, float)) and label != 'Trang thai':
            cell.number_format = '#,##0'

    # Dieu chinh do rong cot
    for col in range(1, 4):
        sheet.column_dimensions[get_column_letter(col)].width = 20

def create_revenue_detail_sheet(sheet, data, header_font, header_fill, border):
    """Tao sheet doanh thu chi tiet"""

    # Tieu de
    sheet['A1'] = "CHI TIET DOANH THU"
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A2'] = f"Thang: {data['month'] or 'Tat ca'}"
    sheet['A2'].font = Font(bold=True, size=12)

    # Header
    headers = ['STT', 'Ngay hoa don', 'Khach hang', 'Ten san pham', 'Loai nhom', 'Loai kinh',
               'Loai tay nam', 'Bo phan', 'Kich thuoc', 'So luong', 'Don gia', 'Thanh tien']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Du lieu
    row_idx = 5
    stt = 1

    for invoice in data['revenue']:
        invoice_date = invoice.get('invoice_date', '')
        if invoice_date:
            try:
                invoice_date = datetime.fromisoformat(invoice_date.replace('Z', '+00:00')).strftime('%Y-%m-%d')
            except:
                invoice_date = invoice_date.split('T')[0] if 'T' in invoice_date else invoice_date

        customer_name = invoice.get('customer_name', '')

        # Lay items cua invoice nay
        invoice_items = [item for item in data.get('invoice_items', []) if item.get('invoice_id') == invoice.get('id')]

        if not invoice_items:
            # Neu khong co items, tao dong trong
            for col in range(1, len(headers) + 1):
                cell = sheet.cell(row=row_idx, column=col)
                cell.border = border
            row_idx += 1
            stt += 1
            continue

        for item_idx, item in enumerate(invoice_items):
            # Lay ten tu lookup maps
            ten_nhom = data['loainhom_map'].get(str(item.get('id_nhom', '')), item.get('id_nhom', ''))
            ten_kinh = data['loaikinh_map'].get(str(item.get('id_kinh', '')), item.get('id_kinh', ''))
            ten_taynam = data['loaitaynam_map'].get(str(item.get('id_taynam', '')), item.get('id_taynam', ''))
            ten_bophan = data['bophan_map'].get(str(item.get('id_bophan', '')), item.get('id_bophan', ''))

            # Lay ten san pham
            sanpham_key = f"{item.get('id_nhom', '')}_{item.get('id_kinh', '')}_{item.get('id_taynam', '')}_{item.get('id_bophan', '')}"
            tensp = data['sanpham_map'].get(sanpham_key, item.get('sanpham_id', ''))

            row_data = [
                stt if item_idx == 0 else '',
                invoice_date if item_idx == 0 else '',
                customer_name if item_idx == 0 else '',
                tensp,
                ten_nhom,
                ten_kinh,
                ten_taynam,
                ten_bophan,
                f"{item.get('ngang', 0)} x {item.get('cao', 0)} x {item.get('sau', 0)}",
                item.get('so_luong', 0),
                item.get('don_gia', 0),
                item.get('thanh_tien', 0)
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border

                if col_idx in [10, 11, 12]:  # So luong, don gia, thanh tien
                    cell.alignment = Alignment(horizontal='right')
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'

            row_idx += 1

        stt += 1

    # Tong ket
    total_row = row_idx + 1
    sheet.cell(row=total_row, column=11).value = "TONG DOANH THU:"
    sheet.cell(row=total_row, column=11).font = Font(bold=True)
    sheet.cell(row=total_row, column=12).value = sum(inv.get('total_amount', 0) for inv in data['revenue'])
    sheet.cell(row=total_row, column=12).font = Font(bold=True)
    sheet.cell(row=total_row, column=12).number_format = '#,##0'

    # Dieu chinh do rong cot
    column_widths = [8, 15, 25, 30, 20, 20, 20, 15, 20, 12, 15, 15]
    for col, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width

def create_expense_detail_sheet(sheet, data, header_font, header_fill, border):
    """Tao sheet chi phi chi tiet"""

    # Tieu de
    sheet['A1'] = "CHI TIET CHI PHI"
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A2'] = f"Thang: {data['month'] or 'Tat ca'}"
    sheet['A2'].font = Font(bold=True, size=12)

    # Header
    headers = ['STT', 'Ngay chi phi', 'Loai chi phi', 'Mo ta', 'So tien (VND)', 'Ty le (%)']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Du lieu
    row_idx = 5
    total_expenses = 0

    for idx, expense in enumerate(data['expenses'], 1):
        expense_date = expense.get('created_at', '')
        if expense_date:
            try:
                expense_date = datetime.fromisoformat(expense_date.replace('Z', '+00:00')).strftime('%Y-%m-%d')
            except:
                expense_date = expense_date.split('T')[0] if 'T' in expense_date else expense_date

        category_name = 'N/A'
        if expense.get('id_lcp') and expense.get('id_lcp') in data['expense_categories']:
            category_name = data['expense_categories'][expense['id_lcp']].get('tenchiphi', 'N/A')

        amount = expense.get('giathanh', 0)
        total_expenses += amount
        ratio = expense.get('ti_le', 0)

        row_data = [
            idx,
            expense_date,
            category_name,
            expense.get('mo_ta', ''),
            amount,
            f"{ratio:.2f}" if ratio else '0.00'
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border

            if col_idx in [5, 6]:  # So tien va ty le
                cell.alignment = Alignment(horizontal='right')
                if col_idx == 5 and isinstance(value, (int, float)):  # So tien
                    cell.number_format = '#,##0'

        row_idx += 1

    # Tong ket
    total_row = row_idx + 1
    sheet.cell(row=total_row, column=5).value = "TONG CHI PHI:"
    sheet.cell(row=total_row, column=5).font = Font(bold=True)
    sheet.cell(row=total_row, column=6).value = total_expenses
    sheet.cell(row=total_row, column=6).font = Font(bold=True)
    sheet.cell(row=total_row, column=6).number_format = '#,##0'

    # Dieu chinh do rong cot
    column_widths = [8, 15, 25, 40, 20, 15]
    for col, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width

def create_payroll_detail_sheet(sheet, data, header_font, header_fill, border):
    """Tao sheet chi phi nhan su chi tiet"""

    # Tieu de
    sheet['A1'] = "CHI TIET CHI PHI NHAN SU"
    sheet['A1'].font = Font(bold=True, size=16)
    sheet['A2'] = f"Thang: {data['month'] or 'Tat ca'}"
    sheet['A2'].font = Font(bold=True, size=12)

    # Header
    headers = ['STT', 'Ma NV', 'Ho ten', 'Ky tinh luong', 'Tong thu nhap', 'Tong khau tru', 'Luong thuc nhan', 'Trang thai']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Du lieu
    row_idx = 5
    total_payroll = 0

    for idx, payroll in enumerate(data['payroll'], 1):
        ma_nv = payroll.get('ma_nv', '')
        employee = data['employees'].get(ma_nv, {})
        ho_ten = employee.get('ho_ten', f'NV {ma_nv}')

        ky_tinh_luong = payroll.get('ky_tinh_luong', '')
        tong_thu_nhap = payroll.get('tong_thu_nhap', 0)
        tong_khau_tru = payroll.get('tong_khau_tru', 0)
        luong_thuc_nhan = payroll.get('luong_thuc_nhan', 0)
        trang_thai = payroll.get('trang_thai', 'draft')

        total_payroll += luong_thuc_nhan

        row_data = [
            idx,
            ma_nv,
            ho_ten,
            ky_tinh_luong,
            tong_thu_nhap,
            tong_khau_tru,
            luong_thuc_nhan,
            trang_thai
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border

            if col_idx in [5, 6, 7]:  # Cac cot tien
                cell.alignment = Alignment(horizontal='right')
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'

        row_idx += 1

    # Tong ket
    total_row = row_idx + 1
    sheet.cell(row=total_row, column=6).value = "TONG CHI PHI NHAN SU:"
    sheet.cell(row=total_row, column=6).font = Font(bold=True)
    sheet.cell(row=total_row, column=7).value = total_payroll
    sheet.cell(row=total_row, column=7).font = Font(bold=True)
    sheet.cell(row=total_row, column=7).number_format = '#,##0'

    # Thong ke them
    stats_row = total_row + 2
    sheet.cell(row=stats_row, column=1).value = "THONG KE:"
    sheet.cell(row=stats_row, column=1).font = Font(bold=True, size=12)

    stats_data = [
        ['Tong so nhan vien co luong', len(data['payroll'])],
        ['Luong trung binh', total_payroll / len(data['payroll']) if data['payroll'] else 0],
        ['Luong cao nhat', max((p.get('luong_thuc_nhan', 0) for p in data['payroll']), default=0)],
        ['Luong thap nhat', min((p.get('luong_thuc_nhan', 0) for p in data['payroll']), default=0)],
    ]

    for row_idx, (label, value) in enumerate(stats_data, stats_row + 1):
        sheet.cell(row=row_idx, column=1).value = label
        sheet.cell(row=row_idx, column=1).border = border

        cell = sheet.cell(row=row_idx, column=2)
        cell.value = value
        cell.border = border
        if isinstance(value, (int, float)):
            cell.number_format = '#,##0'

    # Dieu chinh do rong cot
    column_widths = [8, 12, 25, 15, 18, 18, 18, 12]
    for col, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width

def main():
    parser = argparse.ArgumentParser(description='Xuat bao cao loi nhuan ra Excel')
    parser.add_argument('--month', type=str, help='Thang can xuat (YYYY-MM)', default=None)
    args = parser.parse_args()

    print("Bat dau xuat bao cao loi nhuan ra Excel...")

    # Lay du lieu
    data = get_profit_data(args.month)
    if not data:
        print("Khong the lay du lieu")
        sys.exit(1)

    # Tao ten file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    month_str = args.month.replace('-', '') if args.month else 'all'
    filename = f"bao_cao_loi_nhuan_{month_str}_{timestamp}.xlsx"
    output_path = os.path.join(os.path.dirname(__file__), filename)

    # Tao file Excel
    try:
        create_excel_file(data, output_path)
        print(f"SUCCESS: {filename}")
        print(f"File duoc luu tai: {output_path}")

        # Thong ke
        total_revenue = sum(inv.get('total_amount', 0) for inv in data['revenue'])
        total_expenses = sum(exp.get('giathanh', 0) for exp in data['expenses'])
        total_payroll = sum(p.get('luong_thuc_nhan', 0) for p in data['payroll'])
        total_all_expenses = total_expenses + total_payroll
        total_profit = total_revenue - total_all_expenses

        print("Thong ke:")
        print(f"   - Tong doanh thu: {total_revenue:,.0f} VND")
        print(f"   - Tong chi phi van hanh: {total_expenses:,.0f} VND")
        print(f"   - Tong chi phi nhan su: {total_payroll:,.0f} VND")
        print(f"   - Tong chi phi: {total_all_expenses:,.0f} VND")
        print(f"   - Loi nhuan: {total_profit:,.0f} VND")
        print(f"   - So hoa don: {len(data['revenue'])}")
        print(f"   - So chi phi: {len(data['expenses'])}")
        print(f"   - So nhan vien co luong: {len(data['payroll'])}")

    except Exception as e:
        print(f"Loi khi tao file Excel: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
